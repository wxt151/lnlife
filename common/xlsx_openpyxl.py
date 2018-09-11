# -*- coding: utf-8 -*- 
"""
@__author__ :70486 
@file: OpenpyxlExcel.py
@time: 2017/12/9 18:39
@项目名称:operating
http://openpyxl.readthedocs.io/en/latest/ 读取excel
http://blog.csdn.net/tanzuozhev/article/details/76713387 pandas迭代
http://www.cnblogs.com/chaosimple/p/4153083.html 十分钟搞定pandas
http://pandas.pydata.org/pandas-docs/stable/generated/pandas.DataFrame.set_index.html 官网
https://pandas.pydata.org/pandas-docs/stable/index.html官网
https://www.cnblogs.com/en-heng/p/5630849.html 函数解释
https://jingyan.baidu.com/article/36d6ed1f6c54b01bcf488312.html pandas数据合并
"""

import os
import sys
import time

import pandas as pd


# from tools.PymysqlMain import pymysqls
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows




def read_xlsx(file_name,sheet_name):
    """读取xlsx文件，返回工作表sheet"""
    #打开文档，获取sheet
    import os.path
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        if type(sheet_name) is str:
            sheet = wb[sheet_name]
        else:
            raise KeyError("sheet的名称必须为字符串")
    else:
        raise FileNotFoundError('文件不存在！')
    return sheet

def get_sheet_data(min_row = 1, max_row = None,min_col = 1, max_col = None):
    """获取sheet页面的内容
    :param min_row: 起始行
    :param max_row: 结束行
    :param min_col: 起始列
    :param max_col: 最大列
    """
    #获取标题
    content = sheet.iter_rows(min_row = min_row, max_row=max_row,min_col = min_col,max_col=max_col)
    title_list = [] #存储标题内容
    data_list = []  #存储标题以外的内容
    title_flag = True
    for row in content:
        if title_flag:#如果是标题
            for col in row:
                title_list.append(col.value)
            title_flag = False
        else:
            row_list = []
            for col in row:
                row_list.append(col.value)
            data_list.append(row_list)
    return title_list,data_list


def get_rows_value(min_row = 0,max_row = None):
    """获取行数据
    :param min_row: 起始行
    :param max_row: 结束行
    :return 返回数据列表
    """
    # 获取行，sheet.rows为生成器, 不能使用索引，转换成list后可以
    rows = list(sheet.rows)
    rows_range = rows[min_row:max_row]
    rows_data = []
    for row in rows_range:
        for col in row:
            rows_data.append(col.value)
    return rows_data


def get_cols_value(min_col = 0,max_col = None):
    """获取列数据
       :param min_row: 起始行
       :param max_row: 结束行
       :return 返回数据列表
       """
    # 获取列，sheet.columns, 不能使用索引，转换成list后可以
    cols = list(sheet.columns)
    cols_range = cols[min_col:max_col]
    cols_data = []
    for col in cols_range:
        for cell in col:
            cols_data.append(cell.value)
    return cols_data

def get_sheet_tittle():
    """获取sheet的名称"""
    return sheet.title

def get_column_name(col_index = 1):
    """获取列索引为col_index的列名"""
    #   返回某个列的标题名称
    from openpyxl.utils import get_column_letter
    # col_name = get_column_letter(col_index) 返回的是A，而不是表头名称
    col_name = get_sheet_data(min_row = 0, max_row = 1,min_col = col_index, max_col = col_index )
    return col_name[0]

def total_row_columns(total=True):
    if total:
        total = sheet.max_row  # 单行中，列的数据
    else:
        total = sheet.max_column  # 单列中，行的数据
    return total


if __name__ == '__main__':
    file_name = r"E:\wxt\3W数据量.xlsx"
    sheet_name = "10条数据"
    # # 　读取xlsx的方式
    sheet = read_xlsx(file_name,sheet_name)
    # # title_data,row_col_data = get_sheet_data(min_row = 1, max_row = None, max_col = None)
    # # columnLabel = die_angegebene_keys(row_col_data, title_data, keys="买家ID")
    # temp = total_row_columns()
    # print(temp)




    # nishbushi = [['522230', 'oPRSxv_BOyJHxZKVkL22a4wVWwrA', '18977982260'],
    #              ['522224', 'oPRSxv_BOyJHxZKVkL22a4wVWwrA', '15002151911'],
    #              ['522228', 'oPRSxv_BOyJHxZKVkL22a4wVWwrA', '18977982267']]
    # nihao = []
    # for bushishazi in nishbushi:
    #     nihao.append(list(map(lambda x : x,bushishazi)))
    # print(nihao)