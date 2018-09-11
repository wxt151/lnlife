# -*- coding: utf-8 -*-
'''
#写入例子一
from openpyxl import Workbook

wb = Workbook()
# 激活 worksheet
ws = wb.active
# 数据可以直接分配到单元格中
ws['A1'] = 42
# 可以附加行，从第一列开始附加
ws.append([1, 2, 3])
# Python 类型会被自动转换
import datetime
ws['A3'] = datetime.datetime.now().strftime("%Y-%m-%d")
# 保存文件
wb.save("sample.xlsx")
'''
'''
#写入例子二
# workbook相关
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

wb = Workbook()

dest_filename = 'empty_book.xlsx'

ws1 = wb.active
ws1.title = "range names"

for row in range(1, 40):
    ws1.append(range(600))

ws2 = wb.create_sheet(title="Pi")

ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)
wb.save(filename=dest_filename)
'''

#读取例子一
# -*- coding: utf-8 -*-

from openpyxl.reader.excel import load_workbook
import json

# 读取excel 2007文件
wb = load_workbook(filename=r'test_book.xlsx')

# 显示有多少张表
print("Worksheet range(s):", wb.get_named_ranges())#defined_names.definedName
print("Worksheet name(s):", wb.sheetnames()) # get_sheet_names

# 取第一张表
sheetnames = wb.get_sheet_names()
print(sheetnames)
ws = wb.get_sheet_by_name(sheetnames[0])

# 显示表名，表行数，表列数
print ("Work Sheet Titile:", ws.title)
print ("Work Sheet Rows:", ws.max_row)
print ("Work Sheet Cols:", ws.max_column)


# 建立存储数据的字典
data_dic = {}

# 把数据存到字典中
for rx in range(1, ws.max_row + 1):
    temp_list = []
    pid = rx
    w1 = ws.cell(row=rx, column=1).value
    w2 = ws.cell(row=rx, column=2).value
    w3 = ws.cell(row=rx, column=3).value
    w4 = ws.cell(row=rx, column=4).value
    temp_list = [w1, w2, w3, w4]

    data_dic[pid] = temp_list

# 打印字典数据个数
print ('Total:%d' % len(data_dic))
print (json.dumps(data_dic, encoding="UTF-8", ensure_ascii=False))