# -*- coding: utf-8 -*- 
"""
@__author__ :70486 
@file: OpenpyxlExcel.py
@time: 2017/12/9 18:39
@项目名称:operating
http://openpyxl.readthedocs.io/en/latest/ 读取excel
http://blog.csdn.net/tanzuozhev/article/details/76713387 pandas迭代
http://www.cnblogs.com/chaosimple/p/4153083.html 十分钟搞定pandas
http://pandas.pydata.org/pandas-docs/stable/generated/pandas.DataFrame.set_index.html 官网
https://pandas.pydata.org/pandas-docs/stable/index.html官网
https://www.cnblogs.com/en-heng/p/5630849.html 函数解释
https://jingyan.baidu.com/article/36d6ed1f6c54b01bcf488312.html pandas数据合并
"""

import os
import sys
import time

import pandas as pd

# 获取项目路径下的目录
os.chdir('E:\\operating')
# 将项目路径保存
sys.path.append('E:\\operating')

# from tools.PymysqlMain import pymysqls
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

"""
这个class负责读取数据：从excel中读取数据到电脑
"""


class READEXCEL:
    # def __init__(self, FILEPATH, SHEETNAME=1):
    #     """
    #         FILEPATH :需要文件的位置
    #         SHEETNAME　：读取工作薄的页面或者工作薄名
    #     """
    #     # 判断文件是否存在
    #     if os.path.exists(FILEPATH):
    #         self.excel = FILEPATH
    #
    #         # 判断是否为空
    #         self.readexccel_Data(SHEETNAME)
    #     else:
    #         raise FileNotFoundError('文件不存在！')

    def startReadExcel(self, FILEPATH, SHEETNAME=1):
        """
            FILEPATH :需要文件的位置
            SHEETNAME　：读取工作薄的页面或者工作薄名
        """
        # 判断文件是否存在
        if os.path.exists(FILEPATH):
            self.excel = FILEPATH

            # 判断是否为空
            self.readexccel_Data(SHEETNAME)
        else:
            raise FileNotFoundError('文件不存在！')

    def readexccel_Data(self, sheet):
        """
        此处不严谨：如果输入的字符串都为数字那么就会出错
        :param sheet:
        :return:
        """

        # 创建需要操作的文档
        self.workbook = load_workbook(filename=self.excel)  # 打开文档

        # 判断是根据数字还是文字进行读取sheet，如果是数字的话必须小于现有的长度
        if type(sheet) in [int] and sheet <= len(self.workbook.sheetnames):
            # elf.workbook.sheetnames 打印工作薄名称
            self.sheetbook = self.workbook[self.workbook.sheetnames[sheet - 1]]

        # 如果是文字
        elif type(sheet) in [str]:
            self.sheetbook = self.workbook[sheet]

        # 长度过长时提示
        elif sheet > len(self.workbook.sheetnames):
            print("sheet索要的位置大于现有的长度")

        # 最后输出
        else:
            print("你输入啥咯.")

    def get_sheet_value(self, _value):
        """
        返回读取到单元格的内容，并已cell形式返回
        :param _value:  需要读取的单元格
        :return:
        """
        # 先判断需要寻找的cell位置。如果为空或者其他类型的就提示
        if type(_value) in [int, str]:
            content = self.sheetbook[_value]
            # 打印指定的内容:ws['A4']返回的是一个cell，通过value来获取值
            # content = self.sheetbook[_value].value
            return content
        else:
            print('1级错误')

    def position_sheet_row_value(self, min_row=1, max_row=None, max_col=None):
        """
        指定行读取整行的数据数据信息，数据已经通过value转换了
        :param min_row:  最小的行
        :param max_row:  最大的行
        :param max_col:  最大的列
        :return:
        """
        content = self.sheetbook.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col)
        row_col_data = []  # 存储除了标题以外的内容
        title_data = []  # 只存储标题内容
        _data = True  # 用来控制第一行打印的数据为用例标题
        for row in content:  # 工作薄的全部内容
            row_data = []
            print("***")
            for single in row:  # 遍历每行的数据信息
                if _data:
                    title_data.append(single.value)  # 添加标题
                else:
                    row_data.append(single.value)  # 添加内容
                    print(type(single.value))
            print("***")
            if not _data:
                row_col_data.append(row_data)
                _data = False
        return row_col_data, title_data

    # def die_angegebene_keys(self, row_col_data, title_data, keys="函数"):
    #     """
    #     # 根据指定的keys值执行读取
    #     :param row_col_data: 从case中单独分离出的数据信息
    #     :param title_data:  从case中单独获取的title信息
    #     :param keys:  根据title_data中某个key值进行读取数据
    #     :return:
    #     """
    #     columnLabel = []  # 获取指定key的内容用于做序列名
    #     for title in range(len(title_data)):
    #         if title_data[title] == keys:
    #             for rowColExcel in row_col_data:
    #                 columnLabel.append(rowColExcel[title])
    #             break
    #     return columnLabel

    def position_sheet_cols_value(self, min_row=1, max_row=None, max_col=None):
        '''
        指定列来读取整列的数据，数据已经通过value转换了
        :param min_row:  最小的行
        :param max_row:  最大的行
        :param max_col:  最大的列
        :return:
        '''
        content = self.sheetbook.iter_cols(min_row=min_row, max_row=max_row, max_col=max_col)
        return content

    def position_sheet_row(self, min_row=1, max_row=None, max_col=None):
        """
        指定行读取整行的数据数据信息，数据类型为cell
        :param min_row:  最小的行
        :param max_row:  最大的行
        :param max_col:  最大的列
        :return:
        """
        content = self.sheetbook.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col)
        return content

    def position_sheet_cols(self, min_row=1, max_row=None, max_col=None):
        '''
        指定列来读取整列的数据，数据类型为cell
        :param min_row:  最小的行
        :param max_row:  最大的行
        :param max_col:  最大的列
        :return:
        '''
        content = self.sheetbook.iter_cols(min_row=min_row, max_row=max_row, max_col=max_col)
        return content

    def get_sheet_title(self):
        # 工作薄的名称
        return self.sheetbook.title

    def get_column_letter(self, Number=1):
        #   返回某个列的标题名称
        from openpyxl.utils import get_column_letter
        return get_column_letter(Number)

    def replica_worksheet(self):
        #   返回当前工作薄的复制体对象
        copy_sheet = self.workbook.copy_worksheet(self.sheetbook)
        return copy_sheet

    def total_row_columns(self, total=True):
        """
        为真时，以行为一体，每行的数据信息
        为假时，以列为一体，每列的数据信息
        :param total:
        :return:
        """
        if total:
            content = tuple(self.sheetbook.rows)  # 单行中，列的数据
            ''' 打印长度
                row_cell = tuple(word_sheet.rows)
                row_max_row = len(row_cell) 行的长度
                row_max_col = len(row_cell[0]) 列的长度
            '''
        else:
            content = tuple(self.sheetbook.columns)  # 单列中，行的数据
            '''打印长度 
                col_cell = tuple(word_sheet.columns)
                col_max_col = len(col_cell) 行的长度
                col_max_row = len(col_cell[0]) 列的长度
            '''
        return content

    def attribute_template(self, emplate=None):
        """
        将现有xlsx文档保存为xltx模板
        :param emplate: 需要复制为xltx文档的名字
        :return:
        """

        if type(emplate) in [str]:  # 判断输入的内容是否为字符串
            '''
            将现有的表进行复制并保存为模板。。并后缀名为xltx，如果为xls和xlsx在打开的时候出现问题
            1.判断是否字符串，防止传入数字或者其他类型的
            2.切割一下是否含有xltx后缀名，有则说明可以直接有，没有就拿最后一个名字当做xltx文档的名字
            3.如果为其他格式的就输出说明，
            '''
            genericpath = os.path.splitext(emplate)[1]  # 切割文件后缀名

            if genericpath == '.xltx':
                attribute = os.path.split(emplate)[1]  # 切割最后一个文件的名字
                self.workbook.template = True  # 属性设置
                self.workbook.save(attribute)  # 保存后缀名为xltx的文件
                print('The xlsx document is completed by turning the xitx template.')
                return attribute;  # 返回文件名

            elif genericpath == '':
                attribute = os.path.split(emplate)[1] + '.xltx'  # 切割最后一个文件的名字
                self.workbook.template = True  # 属性设置
                self.workbook.save(attribute)  # 保存后缀名为xltx的文件
                print('The xlsx document is completed by turning the xitx template.')
                return attribute;  # 返回文件名

            elif genericpath in ['xls', 'xlsx', 'txt']:
                print('Files that do not support suffixes such as XLS xlsx text')

            else:
                print('The input file suffix name does not conform.')

        elif emplate == None:
            """
            如果没有传入名字，那么就拿当前xlsx文件的名字作为xltx的名字并在前面加上copy
            """
            path = os.path.splitext(os.path.split(self.excel)[1])[0]
            attribute = 'copy_' + path + '.xltx'  # 设置文件的名字
            self.workbook.template = True
            self.workbook.save(attribute)
            print('Xlsx turns xltx file.')
            return attribute

        else:
            print('你丫的文件输入有误')

    def attribute_document(self, template, document):
        """
        将现有xltx模板转成xlsx文档进行保存
        将现有的模板还原成文档或直接将现有的wb另存为。
        保存为xls文件打开的时候会提示错误
        :param template:  xltx文件
        :param document:  需要保存后的文件
        :return:
        """
        if os.path.exists(template):  # 判断文件是否存在
            """
            1.先判断xltx文件是否存在
            2.在判断传入文件名字是否含有xlsx
            """
            if os.path.splitext(template)[1] == '.xltx':  # 判断模板是不是xltx文件
                attribute = os.path.split(document)  # 将文档切割。切成路径和文件两部分

                if attribute[0] == '':  # 判断路径是否为空，为空说明保存跟模板同一个位置

                    if os.path.splitext(attribute[1])[1] == '.xlsx':  # 判断是否文档保存是否为xlsx文件
                        self.workbook.template = False
                        self.workbook.save(document)
                        print('The xitx template turns to the xlsx document.')
                        return document  # 返回文件名
                    else:
                        print(document + ' : Not a File with a suffix xlsx')
                elif os.path.exists(attribute[0]):  # 有路径说明要保存在指定路径下面

                    if os.path.splitext(attribute[1])[1] == '.xlsx':

                        self.workbook.template = False
                        self.workbook.save(document)
                        print('The xitx template turns to the xlsx document.')
                        return document  # 返回文件名
                    else:
                        print(document + ' : Not a File with a suffix xlsx')
                else:
                    print(document + ' : File path does not exist, please try again.')
            else:
                print(template + ' : Not a File with a suffix xltx')
        else:
            print(template + ' : File does not exist')

    def qisahife(self):
        '''
        一些简单的记录方法
        :return:
        '''
        # 　打印工作薄名称
        print(wb.sheetnames)

        print(len(tuple(word_sheet.rows)))  # 返回行的总数量
        print(len(tuple(word_sheet.rows)[0]))  # 返回单行中列的数量

        print(len(tuple(word_sheet.columns)))  # 返回列的总数量
        print(len(tuple(word_sheet.columns)[0]))  # 返回单列中行的数量


class WRITEEXCEL:
    def __init__(self, FILEPATH, SHEETTITLE='title', _INDEX=None):
        """
        :param FILEPATH:  需要操作的文件路径
        :param SHEETNAME: 工作薄名称，默认为title
        """
        # 判断保存文件的位置是否存在
        path = os.path.split(FILEPATH)
        if os.path.exists(path[0]) or path[0] == '':  # 检验文件是指定存储路径还是存储在本路径下
            if os.path.splitext(path[1])[1] == '.xlsx':  # 检验文件是否为xlsx格式的文件

                self.excel = FILEPATH  # 存储文件的路径保存
                self.writeexcel_Data(_TITLE=SHEETTITLE, _INDEX=_INDEX)  # 调用初始化函数，赋值标题
            elif os.path.splitext(path[1])[1] == '.csv':  # 检验文件是否为xlsx格式的文件
                self.excel = FILEPATH  # 存储文件的路径保存
                self.writeexcel_Data(_TITLE=SHEETTITLE, _INDEX=_INDEX)  # 调用初始化函数，赋值标题

            else:
                print(os.path.splitext(path[1])[1] + ' : 读取文件的格式不对')
        else:
            raise FileNotFoundError('文件不存在！')

    def writeexcel_Data(self, _TITLE, _INDEX):
        """
         此处不严谨：如果输入的字符串都为数字那么就会出错
        :param _TITLE:
        :param _INDEX:
        :return:
        """

        # 创建需要操作的文档
        self.workbook = Workbook()
        if type(_INDEX) in [int]:
            self.create_sheet(_TITLE, _INDEX)
        else:
            self.active_sheet(_TITLE)

    def create_sheet(self, _TITLE, _INDEX):
        """
           如果只创建一个工作薄的话不建议这个方式
           例：
           index设置为0时，会自动生成一个名为sheet的工作薄内容为空
           设置为10时，自动生成9个空内容的工作薄，内存消耗大
         """
        self.work_sheet = self.workbook.create_sheet(title=_TITLE, index=_INDEX)

    def active_sheet(self, _TITLE):
        """
        单独创建一个工作薄.创建之后只有_TITLE标题的工作薄
        :param _TITLE:   工作薄的标题
        :return:
        """
        self.work_sheet = self.workbook.active
        self.work_sheet.title = _TITLE

    def save_woek_sheet(self):
        #   保存工作薄
        self.workbook.save(filename=self.excel)

    def content_cell_single(self, single, content):
        """
        根据单个cell进行赋值
        :param single:  位置
        :param content:  内容
        :return:
        """
        self.work_sheet[single] = content

    def content_cell_row_col(self, col, row, value):
        """
            该语句返回当前设置单元格的内容value
           :param col:
           :param row:
           :param value:
           :return:
       """
        date_ex = self.work_sheet.cell(column=col, row=row, value=value)
        return date_ex

    def content_row_append(self, content):
        """
        对一整行直接写入.
        如果文件已经写入内容时，那么就在下一行写入内容
        :param content:  可以为list也可以是单个内容
        :return:
        """
        self.work_sheet.append(content)

    def get_column_letter(self, Number=1):
        """
        返回指定列的标题名字
        例:
             A    B
        1    x    y
        2    z    x
        :param Number:
        :return:  返回例子中A/B
        """
        from openpyxl.utils import get_column_letter
        return "{0}".format(get_column_letter(Number))

    def time_transformation_timeStamp(self, dt):
        # 时间转时间戳
        if dt == None:
            import datetime
            dt = datetime.datetime.now()
        return dt.timestamp()

    def datetime_transformation(self, dt):
        # 时间戳转时间
        return dt.fromtimestamp()

    def datetime_format(self, format="%Y-%m-%d %H:%M:%S"):
        #   获取当前时间，并按照格式进行返回
        import datetime
        return datetime.datetime.now().strftime(format)

    def merge_excel(self, range):
        """
        合并单元格
        :param range: 需要合并的范围
        :return:
        """
        try:
            self.work_sheet.merge_cells(range)
            pass
        except InsufficientCoordinatesException:
            print('Merge error')

    def ummerge_excel(self, range):
        """
        拆除单元格。。。
        注：
        1.range如果并没有合并，那么执行这个语句会报错
        2.原组合单元格的内容为N时，拆分后第一个单元格的内容为N
        :param range: 现已经合并了，需要拆分的单元格
        :return:
        """
        try:
            self.work_sheet.unmerge_cells(range)
            pass
        except InsufficientCoordinatesException:
            print('Break the merge error')

    def row_col_merge_excel(self, start_row=1, start_column=1, end_row=1, end_column=1):
        """
        指定行列之后进行合并
        注：
        1.当start_row和end_row相等时，说明列之间进行合并
        1.当start_column和end_column相等时，说明行之间进行合并
        :param start_row: 开始行
        :param start_column: 开始列
        :param end_row: 结束行
        :param end_column: 结束列
        :return:
        """
        try:
            self.work_sheet.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row,
                                        end_column=end_column)
            pass
        except InsufficientCoordinatesException:
            print('Merge error')

    def row_col_ummerge_excel(self, start_row=1, start_column=1, end_row=1, end_column=1):
        """
        指定行列之后进行拆分
        注：
        1.当start_row和end_row相等时，说明列之间进行拆分
        2.当start_column和end_column相等时，说明行之间进行拆分
        3.如果传入的单元格并没有合并，那么执行这个语句会报错
        4.原组合单元格的内容为N时，拆分后第一个单元格的内容为N
        :param start_row:开始行
        :param start_column:开始列
        :param end_row:结束行
        :param end_column:结束列
        :return:
        """
        try:
            self.work_sheet.unmerge_cells(start_row=start_row, start_column=start_column, end_row=end_row,
                                          end_column=end_column)
            pass
        except InsufficientCoordinatesException:
            print('Break the merge error')

    def insert_picture(self, range, image):
        """
        插入图片
        :param range: 需要插入图片的位置
        :param image: 插入的图片
        :return:
        """
        from openpyxl.drawing.image import Image
        ws[range] = "You should see three logos below"
        img = Image(image)
        ws.add_image(img, range)

    def bar_comparison_diagram(self, from_rows=True,
                               min_col=1,
                               min_row=1,
                               max_col=2,
                               max_row=2,
                               anchor=None):
        """
        将一定范围内容的数据转换成条形图，进行比较。更方便查看
        :param from_rows:  为真时,拿行作为一个系列。为假时拿列作为一个系列进行比较
        :param min_col:  最小列
        :param min_row:  最小行
        :param max_col:  最大列
        :param max_row:  最大行
        :param anchor:   图标显示的开始位置
        :return:
        """
        from openpyxl.chart import BarChart, Reference
        # 制作表格的数据范围
        values = Reference(self.work_sheet,
                           min_col=min_col,
                           min_row=min_row,
                           max_col=max_col,
                           max_row=max_row)

        chart = BarChart()  # 指定表格对象

        # from_rows为真时,以行为一个系列，比较各行中同一列的Neri
        # from_rows为假时,以列为一个系列，比较各列中同一行的内容
        # titles_from_data为真时，根据from_rows设置的拿第一行或者第一列作为标题名
        # titles_from_data为假时，标题显示为系列1/系列2/以此递增
        # 建议不要将titles_from_data设为真
        chart.add_data(values, from_rows=from_rows)

        ws.add_chart(chart, anchor)

    def comment_remarks(self, range, text='Bug', author='dingdong'):
        """
        设置评论内容及作者...
        :param range: 指定范围设置相应的备注
        :param text: 评论内容
        :param author: 作者
        :return:
        """
        # 添加标红注释
        from openpyxl.comments import Comment
        comment = Comment(text=text, author=author)
        ws[range].comment = comment

    def folding_column(self, min_range, max_range, hidden=False):
        """
         折叠柱（轮廓）（将指定 column name 进行折叠）
        :param min_range:  开始的位置
        :param max_range:  结束的位置
        :param hidden:  should the group be hidden on workbook open or not
        :return:
        """

        if type(min_range) in [str] and type(max_range) in [str]:
            """
            1.先判断开始位置是否小于结束位置：思路：字符串长度以及ASCII的大小
            """
            # 算出值的长度
            length_min = len(min_range)
            length_max = len(max_range)

            # 　开始位置必须在结束位置的前面，所以先判断两个字母的长度
            #   长度小于说明ASCII码也一定小于
            if length_min < length_max:
                self.work_sheet.column_dimensions.group(min_range, max_range, hidden=hidden)

            # 长度相等时，判断长度是不是为1.如果是说明不需要将字符串拆分之后进行计算
            elif length_min == length_max and length_min == 1:
                # 算出值的大小ASCII
                number_min = ord(min_range)
                number_max = ord(max_range)

                if number_min < number_max:
                    self.work_sheet.column_dimensions.group(min_range, max_range, hidden=hidden)
                else:
                    print('When the length is 1, the ASCII is greater than the end....')

            elif length_min == length_max and length_min > 1:
                # 算出值的大小ASCII
                number_min = 0
                number_max = 0

                for number in min_range:
                    number_min = number_min + ord(number)

                for number in max_range:
                    number_max = number_max + ord(number)

                if number_min < number_max:
                    self.work_sheet.column_dimensions.group(min_range, max_range, hidden=hidden)
                else:
                    print('When the length is not 1, ASCII is greater than the end....')
            else:
                print('dayle')
        else:
            print('The parameters given must be characters....')

    def worksheet_color(self, color='1072BA'):
        # 设置工作薄标题颜色
        self.work_sheet.sheet_properties.tabColor = color

    def sfewjq(self):
        # 一些简单的方法
        # 设置标题的2种方式
        ws1 = wb.active
        ws1.title = 'range names'

        ws2 = wb.create_sheet(title='Pi', index=3)  # 设置标题并指定sheet的位置
        # 如果单元格内容为时间，那么可以查看设置时间的格式
        # 为数字时，打印数字格式
        ws['A1'].number_format


class PANDASDATA:

    def __init__(self, _data=None):
        """
        接收excle中读取到的数据
        :param _data:  excle数据源
        :return:
        """
        self._data = _data

    def startPandasData(self, _data):
        """
        接收excle中读取到的数据
        :param _data:  excle数据源
        :return:
        """
        self._data = _data
        return self

    def conversion_series(self):
        '''
        将列表的数据进行转换
        :return:
        '''
        series = pd.Series(self._data)
        return series

    def definition_DataFrame(self, index, periods, columns=None):
        '''
        将字典的业内容进行系列化。
        :param index:  字典中的序列号
        :param columns: 字典中的key
        :return:
        例:
                key1   key2
        index1   1     1
        index2   2     2
        '''
        dates = pd.date_range(index, periods=periods)
        # 转换
        return self.dataFrame(dates, columns)

    def dataFrame(self, index=None, columns=None):
        '''
       将字典业已的内容进行系列化。
        :param index:  字典中的序列号
        :param columns: 字典中的key
        :return:
        例:
                key1   key2
        index1   1     1
        index2   2     2
        '''
        # 转换
        df = pd.DataFrame(self._data, index=index, columns=columns)
        return df

    def functionConcat(self, function, *frames):
        '''
        将多个DataFrame数据集合并之后，将其转成excle文档方便进行查看
        :param function:  新创建的excle文件名
        :param frames:  多个DataFrame合并后的数据集
        :return:
        '''
        result = pd.concat(frames, keys=['readdata', 'storage', 'results'])
        result.to_csv(function + ".csv", index=False, encoding="gbk")

    def contentConcat(self, *frames: "多个Dataframes数据"):

        result = pd.concat(frames)
        return result

    def df_conversion(self, df, data_type='itertuples'):
        '''
        df转换成list的方法，然后给excle输入
        :param df: 通过pandas转换出来的df数据
        :param data_type: 指定转换的方法
        :return:
        '''
        """
        data_type是判断你需要那类方式:运行效率
        itertuples > enumerate > iterrows > range(index)/iloc
        """
        if data_type == 'enumerate':
            list_max = []
            for i, row in enumerate(df.values):
                list_data = []
                for r in row:
                    list_data.append(r.value)
                list_max.append(list_data)
                return list_max
        elif data_type == 'iterrows':
            list_max = []
            for i, row in df.iterrows():
                list_data = []
                for r in row:
                    list_data.append(r.value)
                list_max.append(list_data)
                return list_max
        elif data_type == 'itertuples':
            list_max = []
            for row in df.itertuples():
                list_data = []
                for r in range(1, len(row)):
                    list_data.append(row[r].value)
                list_max.append(list_data)
                return list_max
        elif data_type == 'iloc':
            list_max = []
            for number in range(len(df)):
                list_data = []
                row = df.iloc[number]
                for r in range(len(row)):
                    list_data.append(row[r].value)
                list_max.append(list_data)
                return list_max
        else:
            print('你确定自己输入正确了?、、、')

    def zip_col(self, df, index=1, number=None):
        '''
        通过zip方法，直接返回指定列的数据
        :param df:
        :param index:    需要返回的行
        :param number:   需要返回的列
        :return:
        '''
        if number is not None:
            list_max = []
            for row in zip(df, df[index], df[number]):
                list_data = []
                for r in range(1, len(row)):
                    list_data.append(row[r].value)
                list_max.append(list_data)
        else:
            list_max = []
            for row in zip(df, df[index]):
                list_data = []
                for r in range(1, len(row)):
                    list_data.append(row[r].value)
                list_max.append(list_data)

    def iloc_row(self, df, index=1):
        """
        直接获取相应的行数据
        df['c1'].iloc[x].dtype 指定列的内容，并打印数据类型
        :param df: df数据对象
        :param index:  需要获取的列位置
        :return:
        """
        if index < len(df):
            content = df.iloc[index]
            return content
        else:
            print('长度大于了。。。。。')

    def row_index_header(self, df, index=False, header=False):
        '''
        建议都为假。。。
                header1    header2
        index1    1          2
        index2    3          4
        打印之后的数据为:
            [index1 , 1 , 2 ]
            [index2 , 3 , 4 ]
        :param df:
        :param index: 为真时打印index标签的内容
        :param header: 为真时打印headder标签的内容
        :return:
        '''
        if index and header:
            """
            都为真时，说明有首行header内容以及首列标签index
            将标题以及标签的数据去除之后重新返回数据
            """
            list_max = []
            for row in dataframe_to_rows(df, index=index, header=header):
                list_data = []
                for r in range(1, len(row)):
                    list_data.append(row[r].value)
                list_max.append(list_data)
            return list_max
        elif index is True and header is False:
            """
            index为真时，说明首列标签index保留
            需要对header进行处理，header不是我们想要的数据
            """
            list_max = []
            for row in dataframe_to_rows(df, index=index, header=header):
                list_data = []
                for r in range(1, len(row)):
                    list_data.append(row[r].value)
                list_max.append(list_data)
            return list_max
        elif index is False and header:
            """
            header为真时，说明首行header保留
            需要对index进行处理，index不是我们想要的数据
            """
            list_max = []
            for row in dataframe_to_rows(df, index=index, header=header):
                list_data = []
                for r in range(len(row)):
                    list_data.append(row[r].value)
                list_max.append(list_data)
            return list_max
        elif index is False and header is False:
            """
            都为假时，说明首行header和标签index都没有返回这时不需要进行处理操作直接使用
            """
            list_max = []
            for row in dataframe_to_rows(df, index=index, header=header):
                list_data = []
                for r in range(len(row)):
                    list_data.append(row[r])
                list_max.append(list_data)
            return list_max
        else:
            print('bus')

    def qwjfqajf(self):
        df.to_csv("foo.csv", index=False, encoding="gbk")
        print(pd.read_csv("foo.csv", encoding="gbk"))

        df.to_excel("foo.csv", index=False, encoding="gbk")
        print(pd.read_excel("foo.csv", encoding="gbk"))

        read = READEXCEL(r'E:\drivers\CasePlan\CasrScene\BuyersWechat\-----.xlsx')
        max_row = read.total_row_columns()
        listdata = []
        for kk in range(1, len(max_row)):
            num = 0
            dictdata = {}
            for kl in max_row[kk]:
                print(max_row[0][num].value, kl.value)
                dictdata[max_row[0][num].value] = kl
                num = num + 1
            listdata.append(dictdata)
        data = []
        for kk in range(0, len(max_row[0])):
            data.append(max_row[0][kk].value)

        print(listdata[0]["场景"].value)
        pan = PANDASDATA(listdata)
        df = pan.definition_DataFrame('20130101', 2)
        content = df.iloc[0]
        print(content["场景"].value)

        print("**********")
        content = read.position_sheet_row_value(min_row=2)
        content1 = read.position_sheet_row_value(min_row=2)
        print("**********", len(tuple(content1)))
        df = pd.DataFrame(content, columns=data)
        df["序号"][0] = "66"
        print(df)
        lll = df.iloc[0]
        print(lll[1].value)
        df.to_csv("six6.xlsx", index=False, encoding="gbk")

        print("*-**-*-*-*")
        from openpyxl.utils.dataframe import dataframe_to_rows
        ddf = df.copy()
        list_max = []
        print(ddf)
        for row in dataframe_to_rows(ddf, index=False):
            list_data = []
            print(row)
            for r in range(len(row)):
                print(row[r])
                list_data.append(row[r])
            list_max.append(list_data)
        # 修改某个key值先的Neri
        lists = {1, 2, 3}
        overall_ExcelData.loc[:, ('结果')] = lists
        print(excelData.index)  # 打印所有序列名
        print(excelData.columns)  # 打印标题
        print(excelData.iloc[0])  # 获取第0行的数据
        print(excelData.loc["22"])  # 获取标签序号为”22“的内容数据

        # concat将其进行合并操作.keys将其合并后用索引区分来源于不同DataFrame的数据
        frames = [dfebs, df, dfop]
        result = pd.concat(frames, keys=['dfebs', 'df', 'dfop'])

        # 获取‘status’列标签下面内容为‘xxx’的全部data数据内容
        df = df[df['status'].isin(['xxx'])]

        # 将指定行数的内容进行返回
        read._data = pd.read_csv('path', sep=',', engine='python')
        read._data.get_chunk(5)


class OpenExcelPandas(READEXCEL, PANDASDATA):

    def __init__(self, name='', sheet=','):
        """
        关于_date和_title的解释
        读取excel的数据时：
          _date表示的文件的路径
          _title表示的是工作薄的页面


        通过pandas进行转换时：
          _date表示的数据
          _title表示的是工作薄的名称
        :param name:
        :param sheet:
        """
        self._date = name
        self._title = sheet

    def readCaseExcel(self, title='函数'):
        # 创建工作薄workbook对象
        self.startReadExcel(self._date, self._title) if self._title else self.startReadExcel(self._date)

        # 将case中内容部分的数据（除标题以外的数据）读出
        # 将case中标题的全部内容读出
        self._data, self._title = self.position_sheet_row_value()
        # 获取指定标题的内容
        # columnLabel = self.die_angegebene_keys(row_col_data=self._date, title_data=self._title, keys=title)

        # 通过pandas将数据进行转换
        return self.conversionPandas(title)

    def internal_pandas_read(self, title='函数'):
        genericpath = os.path.splitext(self._date)[1]  # 切割文件后缀名
        if 'xlsx' in genericpath:
            return self.internal_read_excel(title)
        elif 'csv' in genericpath:
            return self.internal_read_csv(title)
        else:
            print("If pandas read the document information, they should pass in the formatted document..")

    def internal_read_excel(self, title="函数"):
        '''
        利用pandas内置函数，直接读取xlsx的数据信息
        并将函数名提取出来，用于序列号的赋值
        :return:
        '''
        self._data = pd.read_excel(self._date, self._title)
        columnLabel = list(self._data[title])  # 设置序列号的名字
        return self.conversion_column(self._data, columnLabel)

    def internal_read_csv(self, title="函数"):
        '''
        读取csv文档数据。sep 为csv切割符号
        header 指定行为矩阵的key
        engine 最好写吧，不然容易意不意外惊不惊喜
        :param title:  矩阵中，拿来设置序列号的相应行key
        :return:
        '''
        self._data = pd.read_csv(self._date, sep=self._title, header=0, engine='python')
        columnLabel = list(self._data[title])  # 设置序列号的名字
        return self.conversion_column(self._data, columnLabel)

    def conversion_column(self, df, columnLabel):
        if columnLabel != None:
            df = df.set_index([columnLabel])  # 设置df数据中的序列号
        return df.fillna(value='')

    def conversionPandas(self, title="函数"):
        '''
        通过已读取的
        :param title:
        :return:
        '''

        self._data = self.dataFrame(columns=self._title)  # 设置标题名
        columnLabel = list(self._data[title])  # 设置序列号的名字

        return self.conversion_column(self._data, columnLabel)


def sql_de_zhixing():
    # `type` tinyint(1) unsigned NOT NULL DEFAULT '0' COMMENT '类型 1买家 2商家 3配送员 4配送中心 5供应商',
    status_number = '0'
    stop_number = '100000'
    sql = """
        SELECT
    	u.id AS '买家ID',
    	u.open_id AS 'open_id',
    	u.phone AS '认证手机'
    FROM
    	lnsm_user AS u
    LEFT JOIN lnsm_buyer AS b ON u.id = b.buyer_id
    WHERE u.type = 1
    ORDER BY `id` DESC
    LIMIT %s, %s
    ;
        """ % (status_number, stop_number)

    pm = pymysqls()
    pm.connects_readModel()
    result = pm.total_vertical_selects(sql)
    pm.closes()
    df = pd.DataFrame(result, columns=['买家ID', 'open_id', '认证手机'])
    nicename = time.strftime('%H-%M-%S', time.localtime())
    path_kkk = r'F:\desktop\数据量%s-%s.csv' % (stop_number, nicename)
    try:
        df.to_csv(path_kkk, index=False, encoding="gbk")
        print("文件所在的位置 %s" % path_kkk)
        print("新建的文件大小:%.2fM" % (os.path.getsize(path_kkk) / 1048576))
    except Exception as e:
        print(e)


if __name__ == '__main__':
    file_path = 'F:\\desktop\\'
    read_name = "买家记录300.xlsx"
    read_name2 = "买家记录300.csv"
    # 　读取xlsx的方式
    # read = OpenExcelPandas(file_path + read_name, sheet='买家记录300')
    # df_excelData = read.internal_pandas_read("买家ID")
    # print(df_excelData)
    print("--------------------")
    # 读取csv的方式
    # read = OpenExcelPandas(file_path + read_name2, sheet=',')
    # df_excelData = read.internal_pandas_read("买家ID")
    # print(df_excelData)
    print("--------------------")
    # 不通过pandas来读取文档
    read = OpenExcelPandas(file_path + read_name, sheet='买家记录300')
    df_excelData = read.readCaseExcel("买家ID")
    print(df_excelData)
